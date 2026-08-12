import ccxt
import requests
import pandas as pd
import numpy as np
import os
import time
import json
from openpyxl import load_workbook, Workbook
from datetime import datetime, timedelta, timezone

# ==========================================
# -1. AYARLAR (config.json)
# ==========================================
CONFIG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")

DEFAULT_CONFIG = {
    "signal_thresholds": {
        "oi_short_btc": 600.0,
        "price_short_usd": 80.0,
        "oi_mid_pct": 0.8,
        "price_mid_pct": 0.5,
        "mid_window": 8,
        "oi_anchor_pct": 1.0,
        "price_anchor_pct": 0.6
    },
    "funding_thresholds": {
        "extreme_pct": 0.0030
    },
    "adaptive_thresholds": {
        "enabled": False,
        "lookback_days": 7,
        "min_days_required": 3,
        "noise_percentile": 75,
        "short_multiplier": 2.0,
        "mid_multiplier": 2.0,
        "anchor_multiplier": 2.0,
        "min_oi_short_btc": 300.0,
        "min_price_short_usd": 40.0,
        "min_oi_mid_pct": 0.4,
        "min_price_mid_pct": 0.25,
        "min_oi_anchor_pct": 0.5,
        "min_price_anchor_pct": 0.3
    },
    "telegram": {
        "min_interval_minutes": 60
    }
}

def load_config():
    """config.json varsa oradan okur; yoksa (ilk çalıştırma) varsayılan değerlerle
    dosyayı kendisi oluşturur. Eşikleri değiştirmek için artık kod açmana gerek yok —
    sadece config.json içindeki sayıyı değiştirip kaydetmen yeterli (script'i yeniden
    başlattığında yeni değerler devreye girer)."""
    if os.path.exists(CONFIG_PATH):
        try:
            with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
                user_config = json.load(f)
            merged = json.loads(json.dumps(DEFAULT_CONFIG))  # derin kopya
            for section, values in user_config.items():
                if section in merged and isinstance(values, dict):
                    merged[section].update(values)
                else:
                    merged[section] = values
            return merged
        except Exception as e:
            print(f"  ⚠️ config.json okunamadı ({e}), varsayılan ayarlar kullanılıyor.")
            return DEFAULT_CONFIG
    else:
        with open(CONFIG_PATH, 'w', encoding='utf-8') as f:
            json.dump(DEFAULT_CONFIG, f, indent=2, ensure_ascii=False)
        print(f"  ℹ️ config.json bulunamadı, varsayılan ayarlarla oluşturuldu: {CONFIG_PATH}")
        return DEFAULT_CONFIG

CONFIG = load_config()

# ==========================================
# 0. TELEGRAM BİLDİRİMİ
# ==========================================
TELEGRAM_BOT_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN", "")
TELEGRAM_CHAT_ID = os.environ.get("TELEGRAM_CHAT_ID", "")

_LAST_TELEGRAM_SENT = None

def should_send_telegram():
    """Telegram mesajını, config'deki min_interval_minutes'e göre saatte bir
    (varsayılan 60 dk) gönderir. Script her 15dk'da bir snapshot alsa bile,
    Telegram'a spam gitmesin diye bu kontrolden geçer."""
    global _LAST_TELEGRAM_SENT
    min_dk = CONFIG.get('telegram', {}).get('min_interval_minutes', 60)
    now = datetime.now(timezone(timedelta(hours=3)))
    if _LAST_TELEGRAM_SENT is None or (now - _LAST_TELEGRAM_SENT).total_seconds() >= min_dk * 60:
        _LAST_TELEGRAM_SENT = now
        return True
    return False

def send_telegram_message(text, parse_mode="HTML"):
    if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_ID:
        print("  ⚠️ TELEGRAM_BOT_TOKEN / TELEGRAM_CHAT_ID tanımlı değil, mesaj gönderilmedi.")
        return
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
    try:
        resp = requests.post(url, data={
            "chat_id": TELEGRAM_CHAT_ID,
            "text": text,
            "parse_mode": parse_mode
        }, timeout=10)
        if not resp.ok:
            print(f"  ❌ Telegram gönderim hatası: {resp.text}")
    except Exception as e:
        print(f"  ❌ Telegram gönderim hatası: {e}")

def build_telegram_report(failed_borsalar, total_oi, global_funding, cvd_spot, cvd_perp, snap_row):
    lines = []
    if failed_borsalar:
        lines.append(f"⚠️ <b>Bağlantı Sağlanamayan Borsalar:</b> {', '.join(failed_borsalar)}")
        lines.append("")
    lines.append(f"🌍 <b>Toplam OI:</b> {total_oi:,.2f} BTC")
    lines.append(f"💰 <b>Fiyat:</b> ${snap_row['price']:,.2f}")
    lines.append(f"⚖️ <b>Küresel Funding (8s):</b> %{global_funding:+.4f}")
    lines.append(f"📊 <b>Spot CVD:</b> {cvd_spot:+.2f} BTC")
    lines.append(f"📊 <b>Perp CVD:</b> {cvd_perp:+.2f} BTC")
    lines.append("")
    lines.append("🎯 <b>ANLIK SİNYAL DURUMU</b>")
    lines.append(f"Fiyat Durumu   : {snap_row['fiyat_durum']}")
    lines.append(f"OI Durumu      : {snap_row['oi_durum']}")
    lines.append(f"Funding Durumu : {snap_row['funding_durum']}")
    lines.append(f"CVD Durumu     : {snap_row['cvd_durum']}")
    lines.append(f"🧭 GENEL DURUM : {snap_row['genel_durum']}")
    return "\n".join(lines)

# ==========================================
# 1. STANDART CCXT İLE ÇALIŞANLAR
# ==========================================
def get_standard_ccxt_data(exchange_name, symbol):
    try:
        exchange_class = getattr(ccxt, exchange_name)
        exchange = exchange_class({'options': {'defaultType': 'swap'}, 'enableRateLimit': True, 'timeout': 10000})
        
        oi_data = exchange.fetch_open_interest(symbol)
        oi = float(oi_data.get('baseVolume') or oi_data.get('openInterest') or 0)
        
        fund_data = exchange.fetch_funding_rate(symbol)
        funding = float(fund_data.get('fundingRate', 0)) * 100
        
        return oi, funding
    except Exception as e:
        return 0, 0

# ==========================================
# 2. ÖZEL MÜDAHALE GEREKTİRENLER
# ==========================================
def get_custom_bybit_data(category='linear', symbol='BTCUSDT'):
    try:
        bybit = ccxt.bybit({'enableRateLimit': True, 'timeout': 10000})
        resp_oi = bybit.publicGetV5MarketOpenInterest({'category': category, 'symbol': symbol, 'intervalTime': '5min'})
        oi = float(resp_oi['result']['list'][0]['openInterest'])
        
        resp_f = bybit.publicGetV5MarketTickers({'category': category, 'symbol': symbol})
        funding = float(resp_f['result']['list'][0]['fundingRate']) * 100
        
        if category == 'inverse':
            price = float(resp_f['result']['list'][0]['lastPrice'])
            oi = oi / price if price > 0 else 0
            
        return oi, funding
    except:
        return 0, 0

def get_custom_hyperliquid_data():
    try:
        url = "https://api.hyperliquid.xyz/info"
        res = requests.post(url, json={"type": "metaAndAssetCtxs"}, timeout=10).json()
        btc_idx = next(i for i, asset in enumerate(res[0]['universe']) if asset['name'] == 'BTC')
        
        oi = float(res[1][btc_idx]['openInterest'])
        funding = float(res[1][btc_idx]['funding']) * 100
        return oi, funding
    except:
        return 0, 0

def get_hyperliquid_funding_8h_real(coin='BTC'):
    try:
        url = "https://api.hyperliquid.xyz/info"
        end_ms = int(datetime.now(timezone.utc).timestamp() * 1000)
        start_ms = end_ms - 8 * 60 * 60 * 1000
        res = requests.post(url, json={
            "type": "fundingHistory",
            "coin": coin,
            "startTime": start_ms,
            "endTime": end_ms
        }, timeout=10).json()
        if not isinstance(res, list) or len(res) == 0:
            return None
        toplam = sum(float(entry['fundingRate']) for entry in res)
        return toplam * 100
    except Exception:
        return None

def get_custom_binance_usd_data():
    try:
        oi_url = "https://dapi.binance.com/dapi/v1/openInterest?symbol=BTCUSD_PERP"
        oi_resp = requests.get(oi_url, timeout=5).json()
        contracts = float(oi_resp['openInterest'])
        
        fund_url = "https://dapi.binance.com/dapi/v1/premiumIndex?symbol=BTCUSD_PERP"
        fund_resp = requests.get(fund_url, timeout=5).json()
        data = fund_resp[0] if isinstance(fund_resp, list) else fund_resp
        
        funding = float(data['lastFundingRate']) * 100
        price = float(data['markPrice'])
        
        oi_usd = contracts * 100
        oi_btc = oi_usd / price if price > 0 else 0
        return oi_btc, funding
    except Exception as e:
        return 0, 0
    
def get_custom_huobi_data(category='linear'):
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/115.0.0.0 Safari/537.36'}
    def _extract_rate(fund_data):
        if isinstance(fund_data, list):
            fund_data = fund_data[0] if fund_data else {}
        if not isinstance(fund_data, dict): return 0.0
        est = fund_data.get('estimated_rate')
        raw = est if est not in (None, '') else fund_data.get('funding_rate', 0)
        try: return float(raw) * 100
        except: return 0.0
    try:
        contract_code = "BTC-USDT" if category == 'linear' else "BTC-USD"
        base = "https://api.hbdm.com/linear-swap-api/v1" if category == 'linear' else "https://api.hbdm.com/swap-api/v1"
        
        oi_resp = requests.get(f"{base}/swap_open_interest?contract_code={contract_code}", headers=headers, timeout=10).json()
        oi_data = oi_resp.get('data', [])
        if not (isinstance(oi_data, list) and len(oi_data) > 0): return 0, 0
        oi_btc = float(oi_data[0].get('amount', 0))

        fund_resp = requests.get(f"{base}/swap_funding_rate?contract_code={contract_code}", headers=headers, timeout=10).json()
        funding = _extract_rate(fund_resp.get('data', {}))
        return oi_btc, funding
    except:
        return 0, 0

def get_binance_cvd(market_type='spot', symbol='BTCUSDT', interval='1h'):

    try:
        now_utc = datetime.now(timezone.utc)
        day_start_utc = now_utc.replace(hour=0, minute=0, second=0, microsecond=0)
        start_ms = int(day_start_utc.timestamp() * 1000)
        end_ms = int(now_utc.timestamp() * 1000)

        url = "https://api.binance.com/api/v3/klines" if market_type == 'spot' else "https://fapi.binance.com/fapi/v1/klines"
        res = requests.get(url, params={
            'symbol': symbol, 'interval': interval,
            'startTime': start_ms, 'endTime': end_ms, 'limit': 1000
        }, timeout=10).json()

        if not isinstance(res, list):
            print(f"  ❌ Binance CVD Hata ({market_type}): beklenmeyen cevap -> {res}")
            return 0.0

        cvd_btc = 0.0
        for candle in res:
            total_vol = float(candle[5])
            taker_buy_vol = float(candle[9])
            cvd_btc += (taker_buy_vol - (total_vol - taker_buy_vol))
        return cvd_btc
    except Exception as e:
        print(f"  ❌ Binance CVD Hata ({market_type}): {e}")
        return 0.0

def fetch_with_retry(fetch_func, *args, retries=2, delay=3, **kwargs):
    """Bir borsa çağrısı geçici bir hatadan (network blip, rate limit, borsanın
    anlık kesintisi vb.) dolayı 0 dönerse, tüm satırı 'başarısız' loglamadan önce
    birkaç kez daha dener. Kalıcı bir sorun varsa (borsa gerçekten çökmüşse)
    yine de son denemeden sonra 0,0 döner ve failed_borsalar listesine düşer."""
    oi, funding = 0, 0
    for attempt in range(retries + 1):
        oi, funding = fetch_func(*args, **kwargs)
        if oi > 0:
            return oi, funding
        if attempt < retries:
            time.sleep(delay)
    return oi, funding

# ==========================================
# 3. KÜRESEL AĞIRLIKLI HESAPLAMA MOTORU
# ==========================================
def get_global_macro_data():
    print(f"\n[{datetime.now().strftime('%H:%M:%S')}] 🌐 USDT ve USD Tahtalarından Makro Veriler Toplanıyor...")
    markets = {
        'Binance_USDT': fetch_with_retry(get_standard_ccxt_data, 'binance', 'BTC/USDT:USDT'),
        'Binance_USD': fetch_with_retry(get_custom_binance_usd_data),
        'Bybit_USDT': fetch_with_retry(get_custom_bybit_data, category='linear', symbol='BTCUSDT'),
        'Bybit_USD': fetch_with_retry(get_custom_bybit_data, category='inverse', symbol='BTCUSD'),
        'OKX_USDT': fetch_with_retry(get_standard_ccxt_data, 'okx', 'BTC/USDT:USDT'),
        'OKX_USD': fetch_with_retry(get_standard_ccxt_data, 'okx', 'BTC/USD:BTC'),
        'Huobi_USDT': fetch_with_retry(get_custom_huobi_data, category='linear'),
        'Huobi_USD': fetch_with_retry(get_custom_huobi_data, category='inverse'),
        'Hyperliquid': fetch_with_retry(get_custom_hyperliquid_data)
    }
    
    hl_funding_8h_real = get_hyperliquid_funding_8h_real('BTC')

    if markets['Binance_USDT'][0] == 0:
        try:
            resp = ccxt.binance({'enableRateLimit': True, 'timeout': 10000}).fapiPublicGetOpenInterest({'symbol': 'BTCUSDT'})
            markets['Binance_USDT'] = (float(resp.get('openInterest', 0)), get_standard_ccxt_data('binance', 'BTC/USDT:USDT')[1])
        except: pass

    FUNDING_INTERVAL_HOURS = {k: 8 for k in markets.keys()}
    FUNDING_INTERVAL_HOURS['Hyperliquid'] = 1

    # Silinen Borsa Loglarını Geri Getirdik
    print("\n--- Borsa ve Tahta Kırılımları (Funding 8s'e normalize edilmiş) ---")
    normalized = {}
    failed_borsalar = []
    for borsa, (oi, funding) in markets.items():
        funding_8h = hl_funding_8h_real if (borsa == 'Hyperliquid' and hl_funding_8h_real is not None) else funding * (8 / FUNDING_INTERVAL_HOURS[borsa])
        normalized[borsa] = (oi, funding_8h)
        if oi > 0:
            etiket = "Funding(8s, gerçek)" if (borsa == 'Hyperliquid' and hl_funding_8h_real is not None) else "Funding(8s)"
            print(f"  ✅ {borsa.ljust(15)}: OI = {oi:>10,.2f} BTC  |  {etiket} = %{funding_8h:+.4f}")
        else:
            failed_borsalar.append(borsa)
            print(f"  ❌ {borsa.ljust(15)}: Bağlantı Sağlanamadı / Veri 0")

    total_oi_btc = sum(oi for oi, _ in normalized.values() if oi > 0)
    weighted_funding_sum = sum(oi * funding_8h for oi, funding_8h in normalized.values() if oi > 0)
    global_weighted_funding = (weighted_funding_sum / total_oi_btc) if total_oi_btc > 0 else 0

    print("-" * 60)
    print(f"  🌍 KÜRESEL TOPLAM OI         : {total_oi_btc:,.2f} BTC")
    print(f"  ⚖️ AĞIRLIKLI FONLAMA ORANI (8s): %{global_weighted_funding:+.4f}\n")

    return total_oi_btc, global_weighted_funding, failed_borsalar

# ==========================================
# 4. ZAMAN SERİSİ VE SİNYAL JENERATÖRÜ
# ==========================================
HISTORY_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "oi_funding_history.xlsx")
COLUMNS = ['tarih', 'saat', 'oi_btc', 'oi_usd', 'funding_pct', 'price', 'cvd_spot_btc', 'cvd_perp_btc', 'funding_durum', 'oi_durum', 'fiyat_durum', 'cvd_durum', 'genel_durum']
FUNDING_COL = COLUMNS.index('funding_pct') + 1

def get_btc_price():
    try:
        r = requests.get("https://api.binance.com/api/v3/ticker/price?symbol=BTCUSDT", timeout=5).json()
        return float(r['price'])
    except: return 0.0

def load_history(path=HISTORY_FILE):
    if not os.path.isfile(path):
        return pd.DataFrame(columns=COLUMNS + ['timestamp'])
    
    df = pd.read_excel(path, engine='openpyxl')
    for col in COLUMNS:
        if col not in df.columns:
            df[col] = 0.0 if col not in ['funding_durum', 'oi_durum', 'fiyat_durum', 'cvd_durum', 'genel_durum', 'tarih', 'saat'] else "N/A"
            
    df['funding_pct'] = df['funding_pct'].astype(float)
    df['timestamp'] = pd.to_datetime(df['tarih'] + ' ' + df['saat'], format='%d.%m.%Y %H:%M')
    return df.sort_values('timestamp').reset_index(drop=True)

def funding_status(current_funding):
    current_funding = float(current_funding)
    esik = CONFIG['funding_thresholds']['extreme_pct']
    if current_funding > esik:
        return "Aşırı Pozitif"
    elif current_funding > 0.0000:
        return "Pozitif"
    elif current_funding < -esik:
        return "Aşırı Negatif"
    elif current_funding < 0.0000:
        return "Negatif"
    return "Nötr"


def compute_adaptive_thresholds(df):
    """Sabit eşikler yerine, son N GÜNÜN (bugün hariç — gün henüz bitmediği için
    yanıltıcı olur) gerçek OI/fiyat oynaklığına bakarak eşikleri kendiliğinden ayarlar.
    Mantık, elle yaptığımız 'gürültü tabanı' analizinin otomatikleştirilmiş hali:
    her gün için 3 katmanın (kısa/orta/anchor) doğal salınımını ölçer, bunun belirli
    bir yüzdelik dilimini (varsayılan %75) alıp bir güvenlik çarpanıyla (varsayılan 2x)
    çarpar. Piyasa sakinse eşikler otomatik küçülür (daha hassas), hareketliyse büyür
    (spam'e dönmez). Yeterli geçmiş yoksa (min_days_required) None döner, çağıran
    taraf statik config değerlerine düşer."""
    ac = CONFIG.get('adaptive_thresholds', {})
    if not ac.get('enabled', False):
        return None
    if 'tarih' not in df.columns or len(df) == 0:
        return None

    df = df.copy()
    df['_gun'] = pd.to_datetime(df['tarih'], format='%d.%m.%Y', errors='coerce').dt.date
    bugun = datetime.now(timezone(timedelta(hours=3))).date()
    df = df[(df['_gun'].notna()) & (df['_gun'] < bugun)]  # bugünü hariç tut, gün tamamlanmamış

    tum_gunler = sorted(df['_gun'].unique())
    if len(tum_gunler) < ac.get('min_days_required', 3):
        return None

    gunler = tum_gunler[-ac.get('lookback_days', 7):]
    df = df[df['_gun'].isin(gunler)]
    mid_window = CONFIG['signal_thresholds'].get('mid_window', 8)

    short_oi, short_price = [], []
    mid_oi_pct, mid_price_pct = [], []
    anchor_oi_pct, anchor_price_pct = [], []

    for gun in gunler:
        gun_df = df[df['_gun'] == gun].sort_values('saat').reset_index(drop=True)
        if len(gun_df) < 3:
            continue

        # Kısa vade: 3-kayıt pencere aralığı (mutlak)
        oi_roll = gun_df['oi_btc'].rolling(3).apply(lambda x: x.max() - x.min()).dropna()
        price_roll = gun_df['price'].rolling(3).apply(lambda x: x.max() - x.min()).dropna()
        short_oi.extend(oi_roll.tolist())
        short_price.extend(price_roll.tolist())

        # Orta vade: mid_window'luk pencere yüzdesel değişim
        if len(gun_df) >= mid_window:
            for i in range(mid_window, len(gun_df) + 1):
                w = gun_df.iloc[i - mid_window:i]
                ref_oi, ref_price = w['oi_btc'].iloc[0], w['price'].iloc[0]
                if ref_oi and ref_price:
                    mid_oi_pct.append(abs(w['oi_btc'].iloc[-1] - ref_oi) / ref_oi * 100)
                    mid_price_pct.append(abs(w['price'].iloc[-1] - ref_price) / ref_price * 100)

        # Anchor: günün tamamının toplam salınımı (yüzdesel)
        ref_oi, ref_price = gun_df['oi_btc'].iloc[0], gun_df['price'].iloc[0]
        if ref_oi and ref_price:
            anchor_oi_pct.append(abs(gun_df['oi_btc'].iloc[-1] - ref_oi) / ref_oi * 100)
            anchor_price_pct.append(abs(gun_df['price'].iloc[-1] - ref_price) / ref_price * 100)

    def yuzdelik(values, varsayilan):
        if not values:
            return varsayilan
        return float(np.percentile(values, ac.get('noise_percentile', 75)))

    st = CONFIG['signal_thresholds']
    sm, mm, am = ac.get('short_multiplier', 2.0), ac.get('mid_multiplier', 2.0), ac.get('anchor_multiplier', 2.0)

    sonuc = {
        'oi_short_btc': max(yuzdelik(short_oi, st['oi_short_btc']) * sm, ac.get('min_oi_short_btc', 300.0)),
        'price_short_usd': max(yuzdelik(short_price, st['price_short_usd']) * sm, ac.get('min_price_short_usd', 40.0)),
        'oi_mid_pct': max(yuzdelik(mid_oi_pct, st['oi_mid_pct']) * mm, ac.get('min_oi_mid_pct', 0.4)),
        'price_mid_pct': max(yuzdelik(mid_price_pct, st['price_mid_pct']) * mm, ac.get('min_price_mid_pct', 0.25)),
        'mid_window': mid_window,
        'oi_anchor_pct': max(yuzdelik(anchor_oi_pct, st['oi_anchor_pct']) * am, ac.get('min_oi_anchor_pct', 0.5)),
        'price_anchor_pct': max(yuzdelik(anchor_price_pct, st['price_anchor_pct']) * am, ac.get('min_price_anchor_pct', 0.3)),
    }
    return sonuc


def compute_signals(df, current_oi, current_funding, current_price):
    """Üç katmanlı ivme (momentum) sinyali üretir:
    - Kısa vade: son 3 kayda (yaklaşık 45dk) göre ani spike'ları yakalar.
    - Orta vade: son 8 kayda (yaklaşık 2 saat) göre hızlıca gelişen bleed'leri yakalar.
    - Gün başı anchor: bugünün İLK kaydına göre kümülatif kayma — saatlerce yayılan,
      her adımda küçük ama toplamda büyük olan yavaş bleed'leri yakalar (Pine'daki
      "1D" anchor mantığının aynısı).
    Üçü de aynı anda çalışır, herhangi biri tetiklenirse sinyal verilir.
    Not: df tüm geçmiş günleri içerebileceğinden, önce bugünün kayıtlarına filtreleniyor
    — aksi halde 'gün başı' referansı yanlışlıkla dünün verisinden gelebilir.
    Eşikler 11.08.2026 verisinden kalibre edildi; daha fazla veri biriktikçe (özellikle
    volatil günler) yeniden ayarlanmalı."""
    fund_status = funding_status(current_funding)

    today_str = datetime.now(timezone(timedelta(hours=3))).strftime('%d.%m.%Y')
    df_today = df[df['tarih'] == today_str] if ('tarih' in df.columns and len(df) > 0) else df.iloc[0:0]

    adaptif = compute_adaptive_thresholds(df)
    st = adaptif if adaptif else CONFIG['signal_thresholds']

    # Katman 1 — kısa vade: gürültü tabanının (~446 BTC / ~120$, 3-kayıt penceresi) ~2 katı
    OI_THRESHOLD = st['oi_short_btc']
    PRICE_THRESHOLD = st['price_short_usd']
    short_oi_up = short_oi_down = short_price_up = short_price_down = False
    if len(df_today) >= 3:
        last_3 = df_today.tail(3)
        max_oi, min_oi = last_3['oi_btc'].max(), last_3['oi_btc'].min()
        short_oi_up = current_oi > (max_oi + OI_THRESHOLD)
        short_oi_down = current_oi < (min_oi - OI_THRESHOLD)
        max_price, min_price = last_3['price'].max(), last_3['price'].min()
        short_price_up = current_price > (max_price + PRICE_THRESHOLD)
        short_price_down = current_price < (min_price - PRICE_THRESHOLD)

    # Katman 2 — orta vade: ~2 saatlik pencere, yüzdesel
    OI_MID_PCT = st['oi_mid_pct']
    PRICE_MID_PCT = st['price_mid_pct']
    MID_WINDOW = st['mid_window']
    mid_oi_up = mid_oi_down = mid_price_up = mid_price_down = False
    if len(df_today) >= MID_WINDOW:
        window = df_today.tail(MID_WINDOW)
        ref_oi, ref_price = window['oi_btc'].iloc[0], window['price'].iloc[0]
        oi_chg = (current_oi - ref_oi) / ref_oi * 100
        price_chg = (current_price - ref_price) / ref_price * 100
        mid_oi_up = oi_chg > OI_MID_PCT
        mid_oi_down = oi_chg < -OI_MID_PCT
        mid_price_up = price_chg > PRICE_MID_PCT
        mid_price_down = price_chg < -PRICE_MID_PCT

    # Katman 3 — gün başı anchor: yavaş, kalıcı bleed
    OI_ANCHOR_PCT = st['oi_anchor_pct']
    PRICE_ANCHOR_PCT = st['price_anchor_pct']
    anchor_oi_up = anchor_oi_down = anchor_price_up = anchor_price_down = False
    if len(df_today) >= 1:
        anchor_oi, anchor_price = df_today['oi_btc'].iloc[0], df_today['price'].iloc[0]
        oi_chg = (current_oi - anchor_oi) / anchor_oi * 100
        price_chg = (current_price - anchor_price) / anchor_price * 100
        anchor_oi_up = oi_chg > OI_ANCHOR_PCT
        anchor_oi_down = oi_chg < -OI_ANCHOR_PCT
        anchor_price_up = price_chg > PRICE_ANCHOR_PCT
        anchor_price_down = price_chg < -PRICE_ANCHOR_PCT

    if len(df_today) < 3:
        oi_status = "Veri Bekleniyor"
        price_status = "Veri Bekleniyor"
    else:
        if short_oi_up or mid_oi_up or anchor_oi_up: oi_status = "Artıyor"
        elif short_oi_down or mid_oi_down or anchor_oi_down: oi_status = "Düşüyor"
        else: oi_status = "Nötr"

        if short_price_up or mid_price_up or anchor_price_up: price_status = "Artıyor"
        elif short_price_down or mid_price_down or anchor_price_down: price_status = "Düşüyor"
        else: price_status = "Nötr"

    return fund_status, oi_status, price_status

def cvd_durumu(cvd_spot, cvd_perp):
    """Spot ve perp CVD'nin yönünü karşılaştırıp diverjans olup olmadığını etiketler.
    Aynı yöndeyse baskı 'uyumlu' (spot ve kaldıraç aynı tarafta), zıt yöndeyse 'diverjans'
    (biri gerçek talebi, diğeri kaldıraç baskısını gösteriyor demektir)."""
    spot_yon = "Long" if cvd_spot > 0 else ("Short" if cvd_spot < 0 else "Nötr")
    perp_yon = "Long" if cvd_perp > 0 else ("Short" if cvd_perp < 0 else "Nötr")

    if spot_yon == "Nötr" or perp_yon == "Nötr":
        etiket = "Zayıf Sinyal"
    elif spot_yon == perp_yon:
        etiket = "Uyumlu"
    else:
        etiket = "Diverjans"

    return f"Spot {spot_yon} / Perp {perp_yon} ({etiket})"

def genel_durum(fund_status, oi_status, price_status, cvd_spot, cvd_perp):
    """
    Tetikleyici Şartı:
    Piyasa yapıcıların tasfiye (likidasyon) avına çıkması için fonlamanın
    'Nötr Pozitif/Negatif' değil, 'Aşırı' seviyelerde olması gerekir.

    Long ve Short taraf simetrik kurulmuştur:
    - long_trap / short_trap: pozisyon HENÜZ birikiyor, fiyat aleyhte gidiyor (tuzak kuruluyor)
    - long_squeeze / short_squeeze: pozisyon tasfiye OLUYOR, an itibariyle kapanış baskısı var
    """
    fund_positive = (fund_status == "Aşırı Pozitif")
    fund_negative = (fund_status == "Aşırı Negatif")

    # Long taraf: funding Aşırı Pozitif -> longlar baskın/kaldıraçlı
    long_trap = (fund_positive and oi_status == "Artıyor" and price_status == "Düşüyor")
    long_squeeze = (fund_positive and oi_status == "Düşüyor" and price_status == "Düşüyor")

    # Short taraf: funding Aşırı Negatif -> shortlar baskın/kaldıraçlı (long tarafın aynası)
    short_trap = (fund_negative and oi_status == "Artıyor" and price_status == "Artıyor")
    short_squeeze = (fund_negative and oi_status == "Düşüyor" and price_status == "Artıyor")

    if short_squeeze:
        return "Sağlıklı Long (Squeeze + Organik Talep)" if cvd_spot > 0 else "Short Squeeze (Zayıf Temel)"

    if long_squeeze:
        absorption_riski = cvd_spot > 0 and cvd_spot > abs(cvd_perp)
        if absorption_riski:
            return "İşlem Açma (Olası Absorption - Spot Alım Baskın)"
        return "Sağlıklı Short (Squeeze + Organik Satış)" if cvd_spot < 0 else "Long Squeeze (Zayıf Temel)"

    if long_trap:
        absorption_riski = cvd_spot > 0 and cvd_spot > abs(cvd_perp)
        if absorption_riski:
            return "İşlem Açma (Olası Absorption - Spot Alım Baskın)"
        return "Long Trap (Devam Riski)"

    if short_trap:
        dagitim_riski = cvd_spot < 0 and abs(cvd_spot) > abs(cvd_perp)
        if dagitim_riski:
            return "İşlem Açma (Olası Dağıtım - Spot Satış Baskın)"
        return "Short Trap (Devam Riski)"

    # Fonlama Nötr Pozitif veya Nötr Negatif ise herhangi bir tasfiye setup'ı aranmaz
    return "İşlem Açma"

def log_snapshot(oi, funding, price, cvd_spot, cvd_perp, path=HISTORY_FILE):
    now = datetime.now(timezone(timedelta(hours=3))).replace(tzinfo=None)
    oi_usd = oi * price
    funding = float(funding)
    
    df_history = load_history(path)
    fund_status, oi_status, price_status = compute_signals(df_history, oi, funding, price)
    cvd_status = cvd_durumu(cvd_spot, cvd_perp)
    genel = genel_durum(fund_status, oi_status, price_status, cvd_spot, cvd_perp)

    row_data = {
        'tarih': now.strftime('%d.%m.%Y'),
        'saat': now.strftime('%H:%M'),
        'oi_btc': oi,
        'oi_usd': oi_usd,
        'funding_pct': funding,
        'price': price,
        'cvd_spot_btc': cvd_spot,
        'cvd_perp_btc': cvd_perp,
        'funding_durum': fund_status,
        'oi_durum': oi_status,
        'fiyat_durum': price_status,
        'cvd_durum': cvd_status,
        'genel_durum': genel
    }
    
    # Yeni satırı mevcut dataframe'e ekle ve tamamını Excel'e bas (Başlık sorunu yaşanmaması için pandas kullanıyoruz)
    new_row_df = pd.DataFrame([row_data])
    
    # Yalnızca timestamp hariç sütunları kaydet
    save_cols = [c for c in df_history.columns if c != 'timestamp']
    df_to_save = pd.concat([df_history[save_cols], new_row_df], ignore_index=True)
    
    # Excel'e yaz
    df_to_save.to_excel(path, index=False, engine='openpyxl')
    wb = load_workbook(path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, min_col=FUNDING_COL, max_col=FUNDING_COL, max_row=ws.max_row):
        row[0].number_format = '0.0000'
    wb.save(path)
    wb.close()
    
    print(f"\n🎯 ANLIK SİNYAL DURUMU")
    print(f"  Fiyat Durumu   : {price_status}")
    print(f"  OI Durumu      : {oi_status}")
    print(f"  Funding Durumu : {fund_status}")
    print(f"  CVD Durumu     : {cvd_status}")
    print(f"  🧭 GENEL DURUM : {genel}")
    
    return pd.DataFrame([row_data])

def compute_trend(df, hours):
    if df.empty: return None
    cutoff = df['timestamp'].max() - pd.Timedelta(hours=hours)
    past = df[df['timestamp'] <= cutoff]
    if past.empty: return None
    
    ref = past.iloc[-1]
    last = df.iloc[-1]

    oi_change_pct = ((last['oi_btc'] - ref['oi_btc']) / ref['oi_btc'] * 100) if ref['oi_btc'] else None
    funding_change = last['funding_pct'] - ref['funding_pct']
    price_change_pct = ((last['price'] - ref['price']) / ref['price'] * 100) if ref['price'] else None

    return {'window_h': hours, 'oi_change_pct': oi_change_pct, 'funding_change': funding_change, 'price_change_pct': price_change_pct}

def print_trend_report(df):
    print("\n📈 ZAMAN SERİSİ TREND RAPORU")
    print("-" * 60)
    for h in [1, 4, 24]:
        t = compute_trend(df, h)
        if t is None:
            continue
        oi_s = f"{t['oi_change_pct']:+.2f}%" if t['oi_change_pct'] is not None else "N/A"
        fund_s = f"{t['funding_change']:+.4f}"
        price_s = f"{t['price_change_pct']:+.2f}%" if t['price_change_pct'] is not None else "N/A"
        print(f"  Son {h:>2}s  ->  OI: {oi_s:>9}   Funding Δ: {fund_s:>9}   Fiyat: {price_s:>9}")
    print("-" * 60)

def run_snapshot_and_report():
    total_oi, global_funding, failed_borsalar = get_global_macro_data()
    price = get_btc_price()
    
    # CVD: bugün UTC 00:00'dan (TR 03:00) itibaren biriken net alım-satım baskısı
    print(f"[{datetime.now().strftime('%H:%M:%S')}] 🔄 CVD Verileri Hesaplanıyor (Bugün 00:00 UTC'den İtibaren)...")
    cvd_spot = get_binance_cvd('spot', 'BTCUSDT', interval='1h')
    cvd_perp = get_binance_cvd('futures', 'BTCUSDT', interval='1h')
    
    print(f"  📊 Spot CVD (bugün) : {cvd_spot:+.2f} BTC")
    print(f"  📊 Perp CVD (bugün) : {cvd_perp:+.2f} BTC\n")
    
    snap_df = log_snapshot(total_oi, global_funding, price, cvd_spot, cvd_perp)

    report_text = build_telegram_report(failed_borsalar, total_oi, global_funding, cvd_spot, cvd_perp, snap_df.iloc[0])
    if should_send_telegram():
        send_telegram_message(report_text)
    else:
        print("  ⏳ Telegram mesajı bu saat penceresinde zaten gönderildi, atlanıyor.")

    df = load_history()
    print_trend_report(df)
    return df

def run_continuous(interval_minutes=15):
    print(f"Başlatılıyor: Her {interval_minutes} dakikada bir snapshot çalıştırılacak.")
    while True:
        try:
            run_snapshot_and_report()
        except Exception as e:
            print(f"  ❌ Beklenmeyen hata: {e}")
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Sonraki çalışma {interval_minutes} dakikada başlayacak...\n")
        time.sleep(interval_minutes * 60)

if __name__ == "__main__":
    run_continuous(15)